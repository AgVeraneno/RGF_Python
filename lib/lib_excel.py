import sys, os
sys.path.append('../lib/')
from openpyxl import load_workbook
from PyQt4 import QtGui

import lib_GUI

class excel():
    def __init__(self, filepath):
        self.file_path = filepath
        self.workbook = None      # current workbook
        self.worksheet = None     # current worksheet
        self.sheet_list = None    # current worksheet list
    def __enter__(self):
        if self.file_path == None:
            gui_obj = lib_GUI.dialog()
            file_path = gui_obj.fileDialog('Select excel file', dir_path='../')
        else:
            file_path = self.file_path
        if os.path.isfile(file_path):
            self.file_path = file_path
            print('Open workbook: ', file_path)
            self.workbook = load_workbook(file_path,data_only=True)     # load workbook using openpyxl, read in data only
            self.sheet_list = self.workbook.sheetnames           # save sheets in work book to sheet_list
            return self
    def __exit__(self, type, value, traceback):
        print('Close workbook'+self.file_path)
    def readSheet(self, sheet):
        self.worksheet = self.workbook[sheet]
        return self.worksheet.iter_rows()
        
        
        
        
        
        
    def open(self, ini_path='../', input_path=None, sheet=None):
        '''
        get file path
        input_path: assign a path for reading excel file.
        sheet: assign a sheet for open automatically
        '''
        if input_path != None:      # open excel file directly using input path
            file_path = input_path
        else:
            gui_obj = gui.dialog()
            file_path = gui_obj.fileDialog('Select excel file', dir_path=ini_path)
        '''
        load workbook and worksheet if file path is available
        '''
        if os.path.isfile(file_path):
            self.file_path = file_path
            print('Open workbook: ', file_path)
            self.workbook = load_workbook(file_path,data_only=True)     # load workbook using openpyxl, read in data only
            self.sheet_list = self.workbook.get_sheet_names()           # save sheets in work book to sheet_list
            if sheet != None:       # auto assign work sheet if sheet is not None
                self.worksheet = self.workbook.get_sheet_by_name(sheet)
            return True
        else:
            print("Can't open the file:", file_path, ". Invaild file path")
            return False
    def getSheet(self, sheet_name=''):
        self.worksheet = self.workbook.get_sheet_by_name(sheet_name)
    def newSheet(self, sheet_name=''):
        self.worksheet = self.workbook.create_sheet(sheet_name)
    def save(self, filename='output'):
        self.workbook.save('../recipe/'+filename+'.xlsx')
'''
debug entry
'''
if __name__ == '__main__':
    app = QtGui.QApplication(sys.argv)
    test = excel()
    test.open()
    test.newSheet('Sheet')
    test.worksheet['B1'] = '=123456/444'
    test.save()
