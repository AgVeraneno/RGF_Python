'''
excel operation library
Author: Alex Lin
updated: 2018/11/30
'''
## exterior library
import os
from openpyxl import load_workbook
from openpyxl import Workbook

class excel():
    def __init__(self, file=None, readfile=True, sheet_name=''):
        self.file_path = file       # initial file_path
        self.read_mode = readfile   # flag for read or write file
        self.ini_sheet = sheet_name # initial sheet name if in write mode
    def __enter__(self):
        ## get file path ##
        if self.file_path == None and self.read_mode:
            import lib_GUI as gui
            gui_obj = gui.dialog()
            self.file_path = gui_obj.fileDialog('Select excel file', dir_path='../')
        elif self.read_mode == False:
            self.newWorkbook(self.ini_sheet)
            return self
        ## open excel file with valid file path
        if os.path.isfile(self.file_path):
            print('Open: ', self.file_path)
            self.workbook = load_workbook(self.file_path,data_only=True)
            self.sheet_list = self.workbook.sheetnames
            return self
        else:
            raise ValueError("Invalid file: "+self.file_path)
    def __exit__(self, type, value, traceback):
        print('Close: '+self.file_path)
    def newWorkbook(self, sheet_name=''):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        if sheet_name != '': self.worksheet.title = sheet_name
    def newSheet(self, sheet_name=''):
        self.worksheet = self.workbook.create_sheet(sheet_name)
    def save(self):
        self.workbook.save(self.file_path)
    def readSheet(self, sheet):
        self.worksheet = self.workbook[sheet]
        return self.worksheet.iter_rows()